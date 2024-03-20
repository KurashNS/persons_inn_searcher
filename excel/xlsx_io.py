from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows

from dateutil.parser import parse, ParserError

from datetime import datetime

from urllib.parse import urlencode
from itertools import product

import threading

_thread_lock = threading.Lock()


class Person:
	def __init__(self) -> None:
		self.last_name: str = ''
		self.first_name: str = ''
		self.patronymic: str = ''
		self.birthday: str = ''

		self.passport_series: str = ''
		self.passport_number: str = ''

		self.inn: str = ''
		self.inn_search_status: str = ''

	def process_values(self):
		self.last_name = self.last_name.strip().capitalize()
		self.first_name = self.first_name.strip().capitalize()
		self.patronymic = self.patronymic.strip().capitalize()

		if isinstance(self.birthday, datetime):
			self.birthday = self.birthday.strftime(format='%d.%m.%Y')
		elif isinstance(self.birthday, (str, int)):
			date_obj = parse(str(self.birthday))
			self.birthday = date_obj.strftime(format='%d.%m.%Y')
		else:
			raise ParserError('Incorrect birth date format')

		self.passport_series = '0' * (4 - len(self.passport_series)) + self.passport_series
		self.passport_number = '0' * (6 - len(self.passport_number)) + self.passport_number

	@property
	def person_id(self) -> str:
		return f'{self.passport_series} {self.passport_number}'

	@property
	def ogu_form_data(self) -> str:
		return urlencode({
			'last_name': self.last_name,
			'first_name': self.first_name,
			'patronymic': self.patronymic,
			'birthday': self.birthday,
			'document_type': '21',
			'document_value': f'{self.passport_series[:2]} {self.passport_series[2:]} {self.passport_number}'
		})

	@property
	def nalog_ru_form_data(self) -> dict[str, str]:
		search_inn_form = {
			'c': 'find',
			'captcha': '',
			'captchaToken': '',
			'fam': self.last_name,
			'nam': self.first_name,
			'otch': self.patronymic,
			'opt_otch': '1',
			'bdate': self.birthday,
			'doctype': '21',
			'docno': f'{self.passport_series[:2]} {self.passport_series[2:]} {self.passport_number}',
			'docdt': '',
		}
		key_to_pop = 'opt_otch' if self.patronymic else 'otch'
		search_inn_form.pop(key_to_pop)
		return search_inn_form

	@classmethod
	def from_json(cls, person_json: dict[str: str]) -> 'Person':
		attributes_map = {
			'Фамилия': 'last_name',
			'Имя': 'first_name',
			'Отчество': 'patronymic',
			'Дата рождения': 'birthday',
			'Серия': 'passport_series',
			'Номер': 'passport_number',
			'ИНН': 'inn',
			'Статус': 'inn_search_status'
		}

		person = cls()
		for json_attr_name, cls_attr_name in product(person_json, attributes_map):
			if cls_attr_name in json_attr_name:
				setattr(person, attributes_map.get(cls_attr_name), person_json.get(json_attr_name))

		return person

	def to_json(self) -> dict[str, str]:
		return {
			'Фамилия': self.last_name,
			'Имя': self.first_name,
			'Отчество': self.patronymic,
			'Дата рождения': self.birthday,
			'Серия': self.passport_series,
			'Номер': self.passport_number,
			'ИНН': self.inn,
			'Статус': self.inn_search_status
		}


def get_persons_list(input_excel_file: str) -> list[Person]:
	persons_table_workbook = load_workbook(filename=input_excel_file)
	persons_table_sheet = persons_table_workbook.active

	header_row = [cell.value for cell in persons_table_sheet[1]]

	attributes_map = {
		'Фамилия': 'last_name',
		'Имя': 'first_name',
		'Отчество': 'patronymic',
		'Дата рождения': 'birthday',
		'Серия': 'passport_series',
		'Номер': 'passport_number'
	}
	persons_list = []
	for row in persons_table_sheet.iter_rows(min_row=2, values_only=True):
		person = Person()
		for cell_value, col_name in zip(row, header_row):
			if col_name:
				try:
					attr = next(attr for attr in attributes_map if col_name in attr or attr in col_name)
				except StopIteration:
					raise ValueError('Incorrect columns in input excel file')

				setattr(person, attributes_map.get(attr), cell_value if cell_value else '')

		if all(getattr(person, attr) for attr in attributes_map.values() if attr != 'patronymic'):
			if all(isinstance(getattr(person, attr), str) for attr in attributes_map.values() if attr != 'birthday'):
				try:
					person.process_values()
					persons_list.append(person)
				except ParserError:
					pass

	if not persons_list:
		raise ValueError('Incorrect input: no person information')

	return persons_list


def output_results(output_excel_file: str, checked_person: Person) -> None:
	with _thread_lock:
		try:
			wb = load_workbook(filename=output_excel_file)
			ws = wb.active
			header = False
		except FileNotFoundError:
			wb = Workbook()
			ws = wb.create_sheet(title='ИНН')
			header = True

		person_json = checked_person.to_json()
		for row in dataframe_to_rows(df=pd.json_normalize(data=person_json), index=False, header=header):
			ws.append(row)

		wb.save(filename=output_excel_file)


if __name__ == '__main__':
	for pers in get_persons_list('input/persons_table.xlsx'):
		print(pers.to_json())
